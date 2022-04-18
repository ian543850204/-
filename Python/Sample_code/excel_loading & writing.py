# Python excel_Loading
# create by Ian
# 2022/4/18

#Execl相關套件讀取
import openpyxl
from openpyxl import *
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException #work book 
from openpyxl.styles import PatternFill # Excel 底色填充

#讀取 Sample 相對路徑
#path = "input/Sample.xlsx"

#讀取 Sample 絕對路徑    
#path_twbiwebserver = "D:/DPS/DPS_rule_base/dist/input/Sample.xlsx"
    
#先讀取絕對路徑
#try:
#    wb = openpyxl.load_workbook(path_twbiwebserver)
#不行則讀取相對路徑
#except:
#    wb = openpyxl.load_workbook(path)
    
# 創建一個空白活頁簿物件
wb_score = Workbook()
    
#設定複製好的Excel為啟用狀態
ws = wb_score.active

'''--------------寫入標題欄位--------------------'''
#row 列 #column 行 #value 值
ws.cell(row=1,column=1,value='A')
ws.cell(row=1,column=2,value='B')
ws.cell(row=1,column=3,value='C')

'''--------------寫入值--------------------'''

#設置 excel 背景色
#判斷是否有 fail 如有 fail 則改變底色

Color =['ff7f50','000000']

#將底色填充為fille
fille = PatternFill('solid', fgColor=Color[0])

#從 列2 開始
excel_row=2

#迴圈數 可自行調整
for i in range(10):
    ws.cell(row=excel_row,column=1,value='C').fill = fille 
    ws.cell(row=excel_row,column=2,value='D').fill = fille 
    ws.cell(row=excel_row,column=3,value='E').fill = fille 

    excel_row=excel_row+1

#相對路徑存檔
wb_score.save('output/test.xlsx')
